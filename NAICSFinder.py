#!/usr/bin/env python3

import alocal
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook
from openpyxl import load_workbook
import time, sys, re



#Starts up webdriver, logs in to aLocalDevelopment, then specfies the search by Single ZCTA and enters a zipcode
wb = Workbook()
ws = wb.create_sheet("NAICS Data") 

#demandNAICS2 = float(sys.argv[3])
demandNAICS6 = float(sys.argv[3])
zipcode = sys.argv[5]
startingIndustry = sys.argv[4]

print(zipcode)

print(startingIndustry)
driver = webdriver.Firefox()
wait = WebDriverWait(driver, 3)
waitLong = WebDriverWait(driver, 60)

#Starts up webdriver and logs in to alocal
alocal.startUp(driver)
alocal.logIn(sys.argv[1], sys.argv[2], driver)
waitLong.until(EC.url_changes("https://alocaldevelopment.com/pages/zipTabulation"))
waitLong.until(EC.presence_of_all_elements_located((By.XPATH, "//div[@class='ng-input']")))
alocal.searchBy(driver) 
waitLong.until(EC.presence_of_all_elements_located((By.XPATH, "//input[@name='zipCode']")))
inputs = alocal.getInputs(driver)

#Keeps track of amount of cells used for Excel spreadsheet 
c = 0
	
staleNessElems = alocal.findCellsIndustries(driver) 
alocal.enterZipcode(str(zipcode), driver)
alocal.activateSearch(driver)
waitLong.until(EC.presence_of_all_elements_located((By.XPATH, "//mat-cell")))
	
try:
	waitLong.until(EC.staleness_of(staleNessElems[0]))
except:
	print("no element")

industryElems = alocal.findCellsIndustries(driver)
demandElems = alocal.findCellsDemand(driver)

#Dictionaries for storing industry data from NAICS 2 digit and NAICS 6 digit
industriesNaics2 = alocal.makeDictionary(driver)
length = len(industryElems)
	


#This for loop navigates through 5 different menus with increasing NAICS Digits(up to 6) then stores relevant data in a dictionary 
for key2 in industriesNaics2.keys():
	
	if(key2 != startingIndustry and startingIndustry != ''):
		continue
		
	startingIndustry = ''
	industriesNaics6 = {}
	print('key ' + key2 + '\n')
	alocal.naicsSelect(3, inputs)
	waitLong.until(EC.presence_of_all_elements_located((By.XPATH, "//ng-dropdown-panel")))
		
	for index1 in range(0, length):
		alocal.naicsDown(3, inputs)
			
		wait.until(EC.presence_of_all_elements_located((By.XPATH, "//ng-dropdown-panel/div/div/div[@class='ng-option ng-star-inserted ng-option-marked'] | div[@class='ng-option ng-option-marked ng-star-inserted']")))
				
		if(key2 == alocal.stripMarkedString(driver, 3, inputs)):
			alocal.naicsSelect(3, inputs)
			alocal.activateSearch(driver)
			waitLong.until(EC.staleness_of(industryElems[0]))
			time.sleep(2)
				
			
			for index2 in range(0, len(alocal.findCellsIndustries(driver))):
				alocal.selectOption(4, inputs, driver, waitLong)
			
				for index3 in range(0, len(alocal.findCellsIndustries(driver))):
					alocal.selectOption(5, inputs, driver, waitLong)
						
						
					for index4 in range(0, len(alocal.findCellsIndustries(driver))):
						alocal.selectOption(6, inputs, driver, waitLong)
							
						
						#Creates a dictionary for industries in 6 digit NAICS and stores industries with net demand greater than .5 in a dictionary 
						industryElemsNAICS6 = alocal.findCellsIndustries(driver)
						demandElemsNAICS6 = alocal.findCellsDemand(driver)
														
						alocal.naicsSelect(7, inputs)
						waitLong.until(EC.presence_of_all_elements_located((By.XPATH, "//ng-dropdown-panel")))
							
						for index5 in range(0, len(industryElemsNAICS6)):
								
							if(float(demandElemsNAICS6[index5].text) >= demandNAICS6):
									
								for index6 in range(0, len(industryElemsNAICS6) + 1):
									alocal.naicsDown(7, inputs)
										
									try:
										wait.until(EC.presence_of_all_elements_located((By.XPATH, 
										"//ng-dropdown-panel/div/div/div[@class='ng-option ng-star-inserted ng-option-marked']" 
									      + "| //ng-dropdown-panel/div/div/div[@class='ng-option ng-option-marked ng-star-inserted']")))	
									except:
										
										alocal.naicsDown(7, inputs)
											
										wait.until(EC.presence_of_all_elements_located((By.XPATH, "//ng-dropdown-panel/div/div/div[@class='ng-option ng-star-inserted ng-option-marked']" 
																	+ "| //ng-dropdown-panel/div/div/div[@class='ng-option ng-option-marked ng-star-inserted']")))
											
									if(industryElemsNAICS6[index5].text == alocal.stripMarkedString(driver, 7, inputs)):
										
										industriesNaics6[industryElemsNAICS6[index5].text] = [float(demandElemsNAICS6[index5].text),alocal.naicsNumber(driver, 7, inputs)]
											
										print(industryElemsNAICS6[index5].text + ' ' + str(industriesNaics6[industryElemsNAICS6[index5].text]) + '\n')
										break
											
						alocal.naicsEscape(7, inputs)
							
			for key6 in industriesNaics6.keys():
				ws['B' + str(list(industriesNaics6).index(key6) + c + 1)] = key2
				ws['B' + str(list(industriesNaics6).index(key6) + c + 1)] = key6
				ws['C' + str(list(industriesNaics6).index(key6) + c + 1)] = int(zipcode)
				ws['D' + str(list(industriesNaics6).index(key6) + c + 1)] = (industriesNaics6[key6])[1]	
				ws['E' + str(list(industriesNaics6).index(key6) + c + 1)] = (industriesNaics6[key6])[0]

			c += len(industriesNaics6)
			wb.save(zipcode + '.xlsx')
			#Breaks when industry is found and data is extracted from html
			break

driver.find_elements(By.XPATH, "//a[@href='/pages/revenueEmployments']")[0].click()

for i in range(1, c + 1):

	time.sleep(2)
	inputs = alocal.getInputs(driver)
	alocal.enterZipcode(str(ws.cell(i,2).value), driver)

	try:
		alocal.enterNAICSNumber(ws.cell(i,3).value, inputs, wait)
	except:
		ws['E' + str(i)] = "Search Up"
		ws['F' + str(i)] = "Search Up"
		continue
		
	alocal.activateSearch(driver)
	
	try:
		wait.until(EC.presence_of_all_elements_located((By.XPATH, "//mat-cell[@class='p-2 mat-cell cdk-column-AnualRevenueEstimate mat-column-AnualRevenueEstimate ng-star-inserted']")))
	except:
		ws['E' + str(i)] = "No Data"
		ws['F' + str(i)] = "No Data"
		continue
		
	ws['E' + str(i)] = driver.find_element(By.XPATH, "//mat-cell[@class='p-2 mat-cell cdk-column-AnualRevenueEstimate mat-column-AnualRevenueEstimate ng-star-inserted']").text
	
	wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@class='mat-tab-label mat-ripple ng-star-inserted']")))

	tab = driver.find_element(By.XPATH, "//div[@class='mat-tab-label mat-ripple ng-star-inserted']")

	driver.execute_script("arguments[0].click();", tab)
	
	wait.until(EC.presence_of_all_elements_located((By.XPATH, "//mat-cell[@class='p-2 mat-cell cdk-column-EmployementEstimate mat-column-EmployementEstimate ng-star-inserted']")))
	
	print("Employment Num:" + driver.find_element(By.XPATH, "//mat-cell[@class='p-2 mat-cell cdk-column-EmployementEstimate mat-column-EmployementEstimate ng-star-inserted']").text)
	
	time.sleep(2)
	
	ws['F' + str(i)] = driver.find_element(By.XPATH, "//mat-cell[@class='p-2 mat-cell cdk-column-EmployementEstimate mat-column-EmployementEstimate ng-star-inserted']").text
	
	wb.save(zipcode + " Revenue.xlsx")

print("done!")
driver.quit()
